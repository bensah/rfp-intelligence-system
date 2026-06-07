"""One-shot backfill for migration rows the uid-keyed Excel sync missed.

When a Form1 Form_ID is edited after the first import (e.g. a -HHMM/-NNNN suffix
is added), the migration's upsert (keyed on `uid`) can no longer match the
existing DB row, so fields like `search_date` / `submitted_by` /
`submitted_by_email` stay blank on those specific rows.

This script matches `source='migration'` rfp_submissions rows to the Excel
Form1 sheet by NORMALISED TITLE and fills ONLY the still-NULL fields. Idempotent.

    python scripts/backfill_form1_meta.py             # apply
    python scripts/backfill_form1_meta.py --dry-run   # preview only
"""
from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path

from openpyxl import load_workbook

_REPO_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(_REPO_ROOT))
sys.path.insert(0, str(Path(__file__).resolve().parent))  # scripts/ for migrate_excel

from migrate_excel import (  # noqa: E402
    DEFAULT_XLSX, build_col_map, map_form1_row_by_header,
)
from db.supabase_client import get_client, safe_execute  # noqa: E402

# search_date + submitted_by + applicant_role come from Form1; submitted_by_email
# is derived from the users table (the Excel 'Email' column is blank for almost
# every row). applicant_role lives under the "the organisation Role" header (col Q).
_FROM_FORM1 = ("search_date", "submitted_by", "applicant_role")


def _norm(t) -> str:
    return re.sub(r"\s+", " ", str(t or "").strip().lower())


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()
    if not args.xlsx or not Path(args.xlsx).exists():
        sys.exit(f"Excel file not found: {args.xlsx}")

    wb = load_workbook(args.xlsx, data_only=True)
    ws = wb["Form1"]
    col_map = build_col_map(ws)
    by_title: dict[str, dict] = {}
    for rn in range(2, ws.max_row + 1):
        row = [ws.cell(row=rn, column=c).value for c in range(1, ws.max_column + 1)]
        rec = map_form1_row_by_header(row, col_map)
        if not rec:
            continue
        key = _norm(rec.get("opportunity_title"))
        if key and key not in by_title:
            by_title[key] = {
                "search_date": rec.get("search_date"),
                "submitted_by": rec.get("submitted_by"),
                "applicant_role": rec.get("applicant_role"),
                "submitted_by_email": rec.get("submitted_by_email"),  # rarely set
            }
    print(f"Form1: {len(by_title)} titles indexed")

    sb = get_client()
    # Submitter name -> email, from the users table (Excel carries no per-row
    # email, but submitters are team members with accounts).
    users = safe_execute(sb.table("users").select("name,email")).data or []
    email_by_name = {_norm(u.get("name")): u.get("email")
                     for u in users if u.get("name") and u.get("email")}
    print(f"users: {len(email_by_name)} name->email entries")

    db_rows = (
        safe_execute(
            sb.table("rfp_submissions")
            .select("uid,opportunity_title,search_date,submitted_by,applicant_role,submitted_by_email")
            .eq("source", "migration")
        ).data or []
    )
    print(f"\nmigration rows in rfp_submissions: {len(db_rows)}")
    print("  per-row diagnostic (uid | search_date | title-match | patch | title):")
    fixed, unmatched, write_fail = 0, [], []
    for r in db_rows:
        uid = r.get("uid")
        title = r.get("opportunity_title")
        meta = by_title.get(_norm(title))
        patch: dict = {}
        if meta:
            for f in _FROM_FORM1:
                if not r.get(f) and meta.get(f):
                    patch[f] = meta[f]
        if not r.get("submitted_by_email"):
            name = patch.get("submitted_by") or r.get("submitted_by") or (meta or {}).get("submitted_by")
            email = (email_by_name.get(_norm(name)) if name else None) or (meta or {}).get("submitted_by_email")
            if email:
                patch["submitted_by_email"] = email
        flag = "MATCH" if meta else "NO-FORM1-MATCH"
        sd = "set" if r.get("search_date") else "NULL"
        print(f"    {str(uid)[:22]:22} sd={sd:4} [{flag:13}] "
              f"patch={patch or '-'}  :: {str(title)[:42]}")
        if not meta and not r.get("search_date"):
            unmatched.append(uid)
        if patch:
            if not args.dry_run:
                resp = safe_execute(sb.table("rfp_submissions").update(patch).eq("uid", uid))
                if len(getattr(resp, "data", None) or []) == 0:
                    write_fail.append(uid)
                    print("        !! UPDATE wrote 0 rows — likely RLS or uid not found")
            fixed += 1
    print(f"\n{'[dry-run] would update' if args.dry_run else 'updated'} {fixed} row(s).")
    if write_fail:
        print(f"WRITE FAILURES (0 rows affected): {write_fail}")
    if unmatched:
        print(f"unmatched (title not in Form1): {unmatched}")


if __name__ == "__main__":
    main()
