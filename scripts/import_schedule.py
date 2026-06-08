"""Import the Excel 'Schedule' sheet into the app (app_settings.schedule_json).

The sheet lists the weekly Monday check-in calls with Note Taker / RFP
Presenter / Meeting Chair. Header row is auto-detected (the row containing
'Call Date'); columns are matched by keyword so minor header wording changes
don't break it.

    python scripts/import_schedule.py             # apply
    python scripts/import_schedule.py --dry-run   # preview only
"""
from __future__ import annotations

import argparse
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
sys.path.insert(0, str(Path(__file__).resolve().parent))

from migrate_excel import DEFAULT_XLSX  # noqa: E402
from core import schedule  # noqa: E402


def _norm(v) -> str:
    return str(v or "").strip().lower()


def _parse_date(v):
    if isinstance(v, datetime):
        return v.date()
    try:
        return datetime.fromisoformat(str(v)[:10]).date()
    except Exception:
        return None


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()
    if not args.xlsx or not Path(args.xlsx).exists():
        sys.exit(f"Excel file not found: {args.xlsx}")

    wb = load_workbook(args.xlsx, data_only=True)
    if "Schedule" not in wb.sheetnames:
        sys.exit("No 'Schedule' sheet in the workbook.")
    ws = wb["Schedule"]
    rows = [list(r) for r in ws.iter_rows(values_only=True)]

    # Locate the header row + the four columns by keyword.
    hdr_i, cols = None, {}
    for i, row in enumerate(rows):
        cells = [_norm(c) for c in row]
        if any("call date" in c for c in cells):
            hdr_i = i
            for j, c in enumerate(cells):
                if "call date" in c:
                    cols["date"] = j
                elif c == "cid":          # new sequential id column (1,2,3…)
                    cols["cid"] = j
                elif "note" in c:
                    cols["note_taker"] = j
                elif "present" in c:
                    cols["presenter"] = j
                elif "chair" in c:
                    cols["chair"] = j
            break
    if hdr_i is None or "date" not in cols:
        sys.exit("Couldn't find the 'Call Date' header in the Schedule sheet.")

    items: list[dict] = []
    for row in rows[hdr_i + 1:]:
        d = _parse_date(row[cols["date"]]) if cols["date"] < len(row) else None
        if not d:
            continue
        def cell(key):
            j = cols.get(key)
            return (str(row[j]).strip() if j is not None and j < len(row)
                    and row[j] is not None else "")
        items.append({"date": d.isoformat(), "cid": cell("cid"),
                      "note_taker": cell("note_taker"),
                      "presenter": cell("presenter"), "chair": cell("chair")})

    print(f"parsed {len(items)} schedule rows")
    for it in items[:6]:
        print(f"  {it['date']} | note:{it['note_taker'][:20]:20} "
              f"| pres:{it['presenter'][:20]:20} | chair:{it['chair'][:20]}"
              .encode("ascii", "replace").decode("ascii"))
    if args.dry_run:
        print("[dry-run] not saved.")
        return
    schedule.set_schedule(items, updated_by="import_schedule.py")
    print(f"saved {len(items)} entries to app_settings.schedule_json")


if __name__ == "__main__":
    main()
