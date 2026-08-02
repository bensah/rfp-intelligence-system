"""One-time migration of a legacy Excel screener into Supabase.

Reads the workbook at `EXCEL_SOURCE_PATH` (or the first *.xlsx beside
the project root as fallback) — expects sheets: Form1, Schedule,
Meeting_Log, Engagement_Log, Active_Grants_Log, Narrative_Log. Upserts
the records into the corresponding Supabase tables with source='migration'.

Usage:
    python scripts/migrate_excel.py [--dry-run] [--xlsx PATH]
"""
from __future__ import annotations

import argparse
import os
import sys
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any, Iterable, Optional

from openpyxl import load_workbook

# One timestamp per sync run. Stamped onto every synced rfp row's created_at so
# the Records list (ordered by created_at DESC) re-floats the whole Excel-synced
# set to the top on each sync — above the auto-scanned rows, which keep their own
# created_at. (created_at isn't used for RFP analytics, only this ordering.)
_SYNC_TS = datetime.now(timezone.utc).isoformat()

# Force UTF-8 on our output streams. This script prints status with non-ASCII
# glyphs (→, ⚠, ·, —, …). When run as a subprocess (Admin → Sync now), Windows
# defaults stdout/stderr to cp1252, so printing e.g. "adopted→MID" raised
# UnicodeEncodeError ('charmap' codec) and the migration died with exit 1 —
# no success, banner never went green. reconfigure() is a no-op where the
# stream is already UTF-8 or doesn't support it.
for _stream in (sys.stdout, sys.stderr):
    try:
        _stream.reconfigure(encoding="utf-8", errors="replace")
    except (AttributeError, ValueError):
        pass

# Allow running as `python scripts/migrate_excel.py` from repo root.
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from db.supabase_client import get_client  # noqa: E402
from core.scorer import CRITERIA, score_submission  # noqa: E402
from core.review_week import review_week_label  # noqa: E402

# Repo-root *.xlsx fallback — Excel files are gitignored so this is a
# developer-local convenience. Resolves to whichever *.xlsx happens to
# sit beside the project root.
_repo_root = Path(__file__).resolve().parent.parent
_xlsx_candidates = sorted(_repo_root.glob("*.xlsx"))
DEFAULT_XLSX = _xlsx_candidates[0] if _xlsx_candidates else None


# ---------------------------------------------------------------------------
# Cell coercion helpers
# ---------------------------------------------------------------------------
def _txt(v: Any) -> Optional[str]:
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def _date(v: Any) -> Optional[str]:
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    try:
        return datetime.fromisoformat(str(v).split(" ")[0]).date().isoformat()
    except Exception:
        return None


def _ts(v: Any) -> Optional[str]:
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.isoformat()
    if isinstance(v, date):
        return v.isoformat()
    return _txt(v)


def _num(v: Any) -> Optional[float]:
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        s = str(v).replace(",", "").replace("$", "").strip()
        try:
            return float(s)
        except ValueError:
            return None


def _int(v: Any) -> Optional[int]:
    n = _num(v)
    return int(n) if n is not None else None


def _bool_yes(v: Any) -> bool:
    return _txt(v) is not None and str(v).strip().lower() in {"yes", "y", "true", "1"}


def _multi(v: Any) -> Optional[list[str]]:
    s = _txt(v)
    if not s:
        return None
    # Excel multi-select uses ';' separator
    parts = [p.strip() for p in s.split(";") if p.strip()]
    return parts or None


# ---------------------------------------------------------------------------
# Sheet -> row mappers
# ---------------------------------------------------------------------------
import re as _re
import unicodedata as _ud


def _norm_header(s: str) -> str:
    """Normalise header text for case/whitespace/dash-style insensitivity."""
    if s is None:
        return ""
    # Normalize unicode (combining chars), strip BOM
    s = _ud.normalize("NFKC", str(s)).replace("﻿", "")
    # Normalize all dash variants to ASCII hyphen
    s = s.replace("–", "-").replace("—", "-").replace("−", "-")
    # Collapse whitespace + lowercase
    return _re.sub(r"\s+", " ", s.strip().lower())


def build_col_map(ws) -> dict[str, int]:
    """Read header row 1 → {normalised_header: col_index}."""
    out: dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        key = _norm_header(v)
        if key and key not in out:
            out[key] = c
    return out


def map_form1_row_by_header(row: list[Any], col_map: dict[str, int],
                             extra: dict[str, Any] | None = None) -> Optional[dict[str, Any]]:
    """Map a Form1 row by HEADER NAME (resilient to column reordering)."""

    def get(*names: str) -> Any:
        for n in names:
            idx = col_map.get(_norm_header(n))
            if idx and idx - 1 < len(row):
                return row[idx - 1]
        return None

    uid = _txt(get("Form_ID", "Form ID", "Form-ID"))
    title = _txt(get("Opportunity Title", "Title"))
    if not uid or not title:
        return None

    return {
        "uid": uid,
        "form_id": uid,
        "source": "migration",
        # Re-float to the top of the Records list on each sync (see _SYNC_TS).
        "created_at": _SYNC_TS,
        "search_date": _ts(get("Search Date", "Search_Date", "SearchDate",
                                "Date Searched", "Date of Search", "Scan Date")),
        "submitted_by": _txt(get("Submitted By")),
        "submitted_by_email": _txt(get("Email", "Submitted By Email",
                                       "Submitter Email", "Contact Email")),
        "opportunity_title": title,
        "brief_description": _txt(get("Brief Description")),
        "date_posted": _date(get("Date Posted")),
        "funding_agency": _txt(get("Funding Agency", "Funder")),
        "call_geographic_scope": _multi(get("Geographic Scope", "Applicant Country")),
        "call_domain_areas": _multi(get("Program Area")),
        "focus_theme": _txt(get("Focus Theme")),
        "opportunity_link": _txt(get("Opportunity Link")),
        # applicant_role: the source workbook header is "Org Role" (col Q);
        # "Applicant Role"/"Role" are accepted as de-branded aliases. Omitting
        # "Org Role" silently dropped the role on every migration row.
        "applicant_role": _txt(get("Applicant Role", "Role", "Org Role")),
        "lead_applicant": _txt(get("Lead Applicant")),
        "sub_applicant": _txt(get("Sub Applicant")),
        "funding_window": _txt(get("Funding Window")),
        "call_submission_deadline": _date(get("Submission Deadline")),
        "expected_award_date": _date(get("Expected award date", "Expected Award Date")),
        "time_to_award": _txt(get("Time to award", "Time to Award")),
        "call_award_value": _num(get("Estimated Value")),
        "currency": _txt(get("Currency")),
        "project_duration": _int(get("Project Duration (months)", "Project Duration")),
        "submission_format": _txt(get("Submission Format")),
        "feasibility": _txt(get("Feasibility")),
        # Criteria headers — CURRENT Excel names first, legacy (pre-2026-06 rename)
        # kept as fallbacks so older workbooks still import.
        "qualification": _txt(get("MUST 1 - Organisational Qualification",
                                  "MUST 1 - Govt alignment")),
        "strategic_fit": _txt(get("MUST 2 - Strategic fit")),
        "capacity": _txt(get("MUST 3 - Implementation capacity",
                             "MUST 3 - Delivery capacity",
                             "MUST 3 - Implementable scope")),
        "geographic_fit": _txt(get("MUST 4 - Geographic fit", "MUST 4 - Compliant")),
        "cofinancing": _txt(get("MUST 5 - Cofinancing & compliance",
                                "MUST 5 - Cofinancing requirements",
                                "MUST 5 - Resourcing / timeline")),
        "funding_quality": _txt(get("PREFER 6 - Funding quality")),
        "funder_relationship": _txt(get("PREFER 7 - Funder relationship",
                                        "PREFER 7 - Monitorable results")),
        "competitiveness": _txt(get("PREFER 8 - Competitiveness",
                                    "PREFER 8 - Partnership advantage")),
        "bid_effort": _txt(get("PREFER 9 - Bid effort",
                               "PREFER 9 - Scale & sustainability")),
        "decline_flags_present": _bool_yes(get("Decline flags present?", "Decline flags present")),
        "key_risks": _txt(get("Key risks (one line)", "Key risks")),
        "decision": _txt(get("Bid Decision", "Decision")),
        "decision_date": _date(get("Decision date")),
        "decision_note": _txt(get("Decision rationale (1-2 lines)",
                                  "Decision rationale (2-3 lines)",
                                  "Decision rationale")),
        "stage": _txt(get("Stage")),
        "proposal_lead": _txt(get("Proposal Lead")),
        "contributors": _multi(get("Contributors/Reviewers", "Contributors")),
        "reviewers": _multi(get("Reviewers")),
        "support_roles": _txt(get("Support ( e.g. tech/finance/compliance)",
                                    "Support (e.g. tech/finance/compliance)",
                                    "Support")),
        "progress_status": _txt(get("Progress Status")),
        "amount_requested": _num(get("Amount Requested")),
        "date_completed": _date(get("Date Completed")),
        "submissions": _int(get("Submissions")) or 1,
        "donor_decision": _txt(get("Donor Decision Status", "Donor Decision")),
        "next_action": _txt(get("Next Action")),
        "assigned_to": _txt(get("Assigned To")),
        "remarks": _txt(get("Remarks")),
        "action_deadline": _date(get("Action Deadline")),
        "last_update": _date(get("Last Update")),
        # "Donor Decision Date" is when the donor decided (= our date_of_approval)
        "date_of_approval": _date(get("Donor Decision Date", "Date of Approval")),
        "amount_secured": _num(get("Amount Secured")),
        "currency_secured": _txt(get("Currency Secured")),
        "donor_program_officer": _txt(get("Donor Program Officer")),
        "next_step": _txt(get("Next Step")),
        "kickoff_date": _date(get("Kick-off Date", "Kickoff Date")),
        **(extra or {}),
    }


# Backwards-compat shim for any caller still using the old positional form
def map_form1_row(row, extra=None):
    raise RuntimeError(
        "map_form1_row(row) is deprecated — use map_form1_row_by_header(row, col_map)"
    )


def _enrich_rfp(rec: dict[str, Any]) -> dict[str, Any]:
    """Compute review_week + alignment_score + auto_recommendation."""
    score, rec_decision = score_submission(
        {k: rec.get(k) for k in CRITERIA},
        bool(rec.get("decline_flags_present")),
    )
    rec["alignment_score"] = score
    rec["auto_recommendation"] = rec_decision

    # review_week: prefer the search_date, fall back to form_start_date / submitted_at
    anchor = rec.get("search_date") or rec.get("form_start_date")
    if anchor:
        try:
            d = datetime.fromisoformat(anchor.replace("Z", "+00:00")).date()
            rec["review_week"] = review_week_label(d)
        except Exception:
            pass
    return rec


def find_header_row(ws, *needed_headers: str, max_search: int = 20) -> Optional[int]:
    """Return the row number whose cells contain ALL of `needed_headers`.
    Search up to `max_search` rows. Header matching is case/whitespace/dash
    insensitive (via _norm_header).
    """
    needed = {_norm_header(h) for h in needed_headers}
    for r in range(1, min(ws.max_row + 1, max_search + 1)):
        seen = {
            _norm_header(ws.cell(row=r, column=c).value)
            for c in range(1, ws.max_column + 1)
        }
        if needed.issubset(seen):
            return r
    return None


def build_col_map_at(ws, header_row: int) -> dict[str, int]:
    """Build {normalised_header: col_index} from a specific header row."""
    out: dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=c).value
        key = _norm_header(v)
        if key and key not in out:
            out[key] = c
    return out


import hashlib as _hashlib


def _meeting_external_id(
    meeting_date: str,
    donor_title: str | None,
    rfp_uid: str | None,
) -> str:
    """Stable identifier for a meeting log row.

    Prefers RFPID (the new column = Form_ID / uid in Form1) for identity,
    since it uniquely identifies an opportunity even if the user renames
    Donor_Title in the dropdown. Falls back to donor_title for legacy rows
    that predate the RFPID column. Owner is NOT part of the key — Excel can
    reassign an owner without us treating the row as new (we want to UPDATE
    the owner in place, preserving is_resolved).
    """
    natural_key = (rfp_uid or "").strip()
    if not natural_key:
        natural_key = (donor_title or "").strip().lower()
    key = f"{meeting_date}|{natural_key}"
    return _hashlib.md5(key.encode("utf-8")).hexdigest()[:16]


def map_meeting_row_by_header(row: list[Any], col_map: dict[str, int]) -> Optional[dict[str, Any]]:
    """Map one Meeting_Log row using header names."""
    def get(*names: str) -> Any:
        for n in names:
            idx = col_map.get(_norm_header(n))
            if idx and idx - 1 < len(row):
                return row[idx - 1]
        return None

    mdate = _date(get("Meeting Date"))
    if not mdate:
        return None
    donor = _txt(get("Donor_Title", "Donor Title"))
    # NEW: RFPID column carries the Form_ID / uid (= rfp_submissions.uid).
    rfpid = _txt(get("RFPID", "RFP_ID", "Form_ID", "Form ID", "UID"))
    owner = _txt(get("Owner"))
    # MID = explicit Meeting-log unique id (new Excel column). It's a STABLE
    # key, so editing the date / donor / RFP no longer orphans the row (and
    # its app-side is_resolved decision) — the row is UPDATED in place instead
    # of re-inserted as a fresh unresolved duplicate. Falls back to the derived
    # id for legacy rows that predate the MID column.
    mid = _txt(get("MID", "Meeting Log ID", "Meeting_Log_ID", "MeetingID",
                   "Meeting ID"))
    ext = (mid.strip() if (mid and mid.strip())
           else _meeting_external_id(mdate, donor, rfpid))
    return {
        "meeting_date": mdate,
        "donor_title": donor,
        "rfp_uid": rfpid,  # Linked RFP — was None before this fix
        "remarks": _txt(get("Remarks / Issues", "Remarks/Issues", "Remarks", "Issues")),
        "actions": _txt(get("Actions / Recommendations", "Actions/Recommendations",
                              "Actions", "Recommendations")),
        "owner": owner,
        "deadline": _date(get("Deadline", "Due Date", "Due")),
        "source": "migration",
        "external_id": ext,
        # Old derived key — used ONLY to adopt a pre-MID row into its new MID
        # key on the first MID sync (preserving is_resolved, no duplicate).
        # Stripped before insert (not a DB column).
        "_derived_id": _meeting_external_id(mdate, donor, rfpid),
    }


# Backwards-compat shim — deprecated, use map_meeting_row_by_header
def map_meeting_row(row, extra=None):
    raise RuntimeError(
        "map_meeting_row(row) is deprecated — use map_meeting_row_by_header(row, col_map)"
    )


def map_engagement_row(row: list[Any]) -> Optional[dict[str, Any]]:
    # Engagement_Log columns start at B (col 2): Date, Donor, Type, Format, Lead, Contacts, Purpose, Outcome, LinkedRFP
    def c(i: int) -> Any:
        return row[i - 1] if i - 1 < len(row) else None

    edate = _date(c(2))
    if not edate:
        return None
    donor = _txt(c(3))
    internal_lead = _txt(c(6))
    ext_key = f"{edate}|{(donor or '').strip().lower()}|{(internal_lead or '').strip().lower()}"
    return {
        "engagement_date": edate,
        "donor": donor,
        "engagement_type": _txt(c(4)),
        "format": _txt(c(5)),
        "internal_lead": internal_lead,
        "donor_contacts": _txt(c(7)),
        "purpose": _txt(c(8)),
        "outcome": _txt(c(9)),
        "linked_rfp_uid": _txt(c(10)),
        "source": "migration",
        "external_id": _hashlib.md5(ext_key.encode("utf-8")).hexdigest()[:16],
    }


def map_active_grant_row(row: list[Any]) -> Optional[dict[str, Any]]:
    # Active_Grants_Log columns start at B (col 2)
    def c(i: int) -> Any:
        return row[i - 1] if i - 1 < len(row) else None

    gid = _txt(c(2))
    if not gid:
        return None
    return {
        "grant_id": gid,
        "donor_title": _txt(c(3)),
        "form_id_link": _txt(c(4)),
        "award_date": _date(c(5)),
        "end_date": _date(c(6)),
        "report_type": _txt(c(7)),
        "report_due_date": _date(c(8)),
        "submitted_date": _date(c(9)),
        "status": _txt(c(10)),
        "owner": _txt(c(11)),
        "remarks": _txt(c(13)),
        "source": "migration",
    }


def map_narrative_row(row: list[Any]) -> Optional[dict[str, Any]]:
    # Narrative_Log columns start at B (col 2): VersionDate, Title, UsedIn, UsedWith, DateUsed, Status, Link, Owner
    def c(i: int) -> Any:
        return row[i - 1] if i - 1 < len(row) else None

    vdate = _date(c(2))
    if not vdate:
        return None
    title = _txt(c(3))
    ext_key = f"{vdate}|{(title or '').strip().lower()}"
    return {
        "version_date": vdate,
        "narrative_title": title,
        "used_in": _txt(c(4)),
        "used_with": _txt(c(5)),
        "date_used": _date(c(6)),
        "status": _txt(c(7)),
        "link_location": _txt(c(8)),
        "owner": _txt(c(9)),
        "source": "migration",
        "external_id": _hashlib.md5(ext_key.encode("utf-8")).hexdigest()[:16],
    }


def map_schedule_row(row: list[Any]) -> Optional[dict[str, Any]]:
    # Schedule columns start at B (col 2): CallDate, NoteTaker, Presenter, Chair
    def c(i: int) -> Any:
        return row[i - 1] if i - 1 < len(row) else None

    cdate = _date(c(2))
    if not cdate:
        return None
    return {
        "call_date": cdate,
        "note_taker": _txt(c(3)),
        "rfp_presenter": _txt(c(4)),
        "meeting_orgr": _txt(c(5)),
    }


# ---------------------------------------------------------------------------
# Migration driver
# ---------------------------------------------------------------------------
def _rows(ws, header_row: int) -> Iterable[list[Any]]:
    for r in range(header_row + 1, ws.max_row + 1):
        row = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        if any(v is not None and str(v).strip() != "" for v in row):
            yield row


def _load_workbook_retrying(xlsx_path: Path, attempts: int = 4):
    """Open the workbook, retrying on PermissionError.

    The source file often lives in OneDrive; when OneDrive is mid-sync it
    holds a transient Windows file lock, so a startup auto-sync hits
    PermissionError and only succeeds a few seconds later. Retry with a short
    backoff so the auto-sync rides through the lock instead of failing.
    """
    import time as _time
    last: Exception | None = None
    for i in range(max(1, attempts)):
        try:
            return load_workbook(xlsx_path, data_only=True)
        except PermissionError as exc:
            last = exc
            if i < attempts - 1:
                print(f"  file locked (PermissionError) — retrying in "
                      f"{1.5 * (i + 1):.1f}s ({i + 1}/{attempts})…")
                _time.sleep(1.5 * (i + 1))
    raise last if last else RuntimeError("could not open workbook")


def migrate(xlsx_path: Path, dry_run: bool = False) -> None:
    wb = _load_workbook_retrying(xlsx_path)
    sb = None if dry_run else get_client()

    def upsert(table: str, rows: list[dict[str, Any]], conflict_key: str | None = None) -> None:
        if not rows:
            print(f"  {table}: 0 rows — skipping")
            return
        print(f"  {table}: {len(rows)} rows")
        if dry_run:
            return
        if conflict_key:
            sb.table(table).upsert(rows, on_conflict=conflict_key).execute()
        else:
            sb.table(table).insert(rows).execute()

    def merge_by_external_id(
        table: str,
        rows: list[dict[str, Any]],
        updatable_fields: list[str],
    ) -> None:
        """Merge Excel rows into a table keyed by external_id.

        Existing rows (source='migration') get `updatable_fields` UPDATED.
        Any other columns (e.g. app-managed toggles) are PRESERVED.
        Rows whose external_id is new get INSERTED.
        """
        if not rows:
            print(f"  {table}: 0 rows — skipping")
            return
        # In-Excel dedup safety net
        seen: set[str] = set()
        uniq: list[dict[str, Any]] = []
        for r in rows:
            k = r.get("external_id")
            if not k or k in seen:
                continue
            seen.add(k)
            uniq.append(r)
        if len(uniq) < len(rows):
            print(f"  {table}: deduped {len(rows) - len(uniq)} same-key rows in Excel")
        rows = uniq
        print(f"  {table}: {len(rows)} rows mapped")
        if dry_run:
            return
        existing = (
            sb.table(table)
            .select("id,external_id")
            .eq("source", "migration")
            .execute()
            .data
            or []
        )
        ext_to_id = {r["external_id"]: r["id"] for r in existing if r.get("external_id")}
        inserts: list[dict[str, Any]] = []
        updated = 0
        for r in rows:
            ext = r["external_id"]
            if ext in ext_to_id:
                payload = {f: r.get(f) for f in updatable_fields}
                sb.table(table).update(payload).eq("id", ext_to_id[ext]).execute()
                updated += 1
            else:
                inserts.append(r)
        if inserts:
            sb.table(table).insert(inserts).execute()
        print(f"  {table}: {updated} updated · {len(inserts)} inserted")

    # --- rfp_submissions (from Form1, header row 1)
    print("[Form1 -> rfp_submissions]")
    rfp_rows: list[dict[str, Any]] = []
    skipped: list[tuple[int, Any, Any, str]] = []
    ws = wb["Form1"]
    col_map = build_col_map(ws)
    print(f"  {len(col_map)} columns detected by header name")

    for excel_row_num in range(2, ws.max_row + 1):
        row = [ws.cell(row=excel_row_num, column=c).value
               for c in range(1, ws.max_column + 1)]
        if not any(v is not None and str(v).strip() != "" for v in row):
            continue  # blank row
        # Don't pass extra={"submissions": 1} — the function reads Submissions
        # from the row directly (with its own default-to-1 fallback) and
        # **extra here would CLOBBER the real value.
        rec = map_form1_row_by_header(row, col_map)
        if rec:
            rfp_rows.append(_enrich_rfp(rec))
        else:
            form_id_col = col_map.get(_norm_header("Form_ID")) or 2
            title_col = col_map.get(_norm_header("Opportunity Title")) or 5
            uid = row[form_id_col - 1] if form_id_col - 1 < len(row) else None
            title = row[title_col - 1] if title_col - 1 < len(row) else None
            reason = "missing Form_ID" if not _txt(uid) else "missing Opportunity Title"
            skipped.append((excel_row_num, uid, title, reason))
    if skipped:
        print(f"  ⚠ {len(skipped)} row(s) skipped:")
        for r_num, uid, title, reason in skipped:
            print(f"    row {r_num}: uid={uid!r} title={title!r} → {reason}")
    # Fill submitted_by_email from the users table — the Excel 'Email' column is
    # blank for almost every row, but submitters are team members with accounts.
    # Done before upsert so a re-migration never clobbers an email with a blank.
    if not dry_run and sb is not None:
        try:
            _users = sb.table("users").select("name,email").execute().data or []
            _email_by_name = {
                (u.get("name") or "").strip().lower(): u.get("email")
                for u in _users if u.get("name") and u.get("email")
            }
            filled = 0
            for _r in rfp_rows:
                if not _r.get("submitted_by_email"):
                    _e = _email_by_name.get((_r.get("submitted_by") or "").strip().lower())
                    if _e:
                        _r["submitted_by_email"] = _e
                        filled += 1
            if filled:
                print(f"  derived submitted_by_email for {filled} row(s) from users")
        except Exception as exc:
            print(f"  (skipped email derivation: {exc})")
    # Excel-as-source-of-truth (2026-06-25), NON-NULL merge. The UID both DEDUPS
    # and DRIVES UPDATES: new Form_IDs are INSERTED; existing migration rows are
    # UPDATED field-by-field, but ONLY where the Excel cell HAS a value — a blank
    # Excel cell PRESERVES whatever the app already stored, so app-side edits and
    # derived values aren't wiped by an empty column. This supersedes the
    # 2026-06-19 insert-only rule (which never propagated Excel edits at all). The
    # old guard existed to stop OLD-scale criteria values landing in the renamed
    # columns during the one-time rename; the workbook now carries current values.
    # Columns NOT in the Excel mapping are never touched.
    if dry_run:
        print(f"  rfp_submissions (dry-run): {len(rfp_rows)} mapped — would insert "
              "new + update existing on uid (non-null cells only; blanks preserved)")
    elif sb is not None and rfp_rows:
        try:
            _ex = sb.table("rfp_submissions").select("uid").execute().data or []
            _existing_uids = {(e.get("uid") or "") for e in _ex}
        except Exception as _e:
            print(f"  ⚠ could not read existing uids ({_e}); skipped rfp sync for safety")
            _existing_uids = None
        if _existing_uids is not None:
            _new = [r for r in rfp_rows
                    if r.get("uid") and r["uid"] not in _existing_uids]
            _existing = [r for r in rfp_rows
                         if r.get("uid") and r["uid"] in _existing_uids]
            for i in range(0, len(_new), 200):
                sb.table("rfp_submissions").insert(_new[i:i + 200]).execute()
            # Update existing rows with ONLY the Excel cells that carry a value;
            # None (blank cell / missing column) is dropped so it can't null out
            # a stored value. created_at (_SYNC_TS) + computed fields are always
            # non-None, so re-float ordering + scores still refresh.
            _upd = 0
            for r in _existing:
                payload = {k: v for k, v in r.items() if v is not None}
                if not payload:
                    continue
                sb.table("rfp_submissions").update(payload).eq("uid", r["uid"]).execute()
                _upd += 1
            # Tombstone only the brand-NEW uids in the permanent seen-ledger so
            # they're remembered (never silently re-scanned in) even if later deleted.
            if _new:
                try:
                    from core import seen_ledger
                    seen_ledger.record(_new, reason="migration")
                except Exception as _e:
                    print(f"  (seen-ledger record skipped: {_e})")
            print(f"  rfp_submissions: {len(_new)} inserted · {_upd} updated from "
                  "Excel (non-null cells; blanks preserved)")

    # --- Seed source_registry from the Opportunity Link column. Adds only NEW
    # hosts (deduped against the existing registry), as status='pending' so Bernard
    # verifies/corrects the actual Host / listing URL manually afterwards.
    print("[Form1 Opportunity Link -> source_registry]")
    if dry_run:
        print("  source_registry (dry-run): would seed new hosts from Opportunity Link")
    elif sb is not None and rfp_rows:
        try:
            from core.source_registry import normalize_host
            existing_hosts = {
                r.get("host") for r in
                (sb.table("source_registry").select("host").execute().data or [])}
            seen_h: set[str] = set()
            seed_rows: list[dict[str, Any]] = []
            for r in rfp_rows:
                link = (r.get("opportunity_link") or "").strip()
                if not link.lower().startswith("http"):
                    continue
                h = normalize_host(link)
                if not h or h in existing_hosts or h in seen_h:
                    continue        # don't add duplicates
                seen_h.add(h)
                seed_rows.append({
                    "host": h, "classification": "unknown", "status": "pending",
                    "detected_as": "migration", "sample_url": link[:600],
                    "sample_title": (r.get("opportunity_title") or "")[:300] or None,
                    "verified_by": "excel-migration",
                })
            for i in range(0, len(seed_rows), 200):
                sb.table("source_registry").upsert(
                    seed_rows[i:i + 200], on_conflict="host").execute()
            print(f"  source_registry: seeded {len(seed_rows)} NEW host(s) "
                  "(pending manual verify of Host / listing URL)")
        except Exception as _e:
            print(f"  (source_registry seed skipped: {_e})")

    # --- meeting_logs (Meeting_Log) — header auto-detected
    print("[Meeting_Log -> meeting_logs]")
    ws_m = wb["Meeting_Log"]
    hr_m = find_header_row(ws_m, "Meeting Date", "Donor_Title", "Owner")
    if hr_m is None:
        print("  ⚠ Could not find header row in Meeting_Log — sheet skipped")
    else:
        print(f"  Header detected at row {hr_m}")
        col_map_m = build_col_map_at(ws_m, hr_m)
        m_rows: list[dict[str, Any]] = []
        for r in range(hr_m + 1, ws_m.max_row + 1):
            row = [ws_m.cell(row=r, column=c).value for c in range(1, ws_m.max_column + 1)]
            if not any(v is not None and str(v).strip() != "" for v in row):
                continue
            rec = map_meeting_row_by_header(row, col_map_m)
            if rec:
                m_rows.append(rec)
        # In-memory dedup safety net: collapse exact-duplicate rows BEFORE insert.
        seen_keys: set[str] = set()
        uniq_m_rows: list[dict[str, Any]] = []
        for r in m_rows:
            key = r["external_id"]
            if key in seen_keys:
                continue
            seen_keys.add(key)
            uniq_m_rows.append(r)
        if len(uniq_m_rows) < len(m_rows):
            print(f"  meeting_logs: deduped {len(m_rows) - len(uniq_m_rows)} "
                  "rows with same external_id within Excel before insert")
        m_rows = uniq_m_rows
        print(f"  meeting_logs: {len(m_rows)} rows mapped")

        # MERGE semantics: existing rows get Excel-managed fields UPDATED;
        # is_resolved (and any app-side toggles) are PRESERVED. New rows
        # are inserted fresh.
        if not dry_run and m_rows:
            existing = (
                sb.table("meeting_logs")
                .select("id,external_id,is_resolved")
                .eq("source", "migration")
                .execute()
                .data
                or []
            )
            # external_id -> LIST of existing rows (a list, not one id) so we
            # can COLLAPSE rows that share a key — leftovers from earlier
            # duplicate-creating sync runs. is_resolved is fetched so the
            # surviving copy is the resolved one when there's a choice.
            ext_to_rows: dict[str, list[dict[str, Any]]] = {}
            for e in existing:
                if e.get("external_id"):
                    ext_to_rows.setdefault(e["external_id"], []).append(e)

            def _survivor(rows: list[dict[str, Any]]):
                """(keep, [extras]) — keep a resolved copy if any so an in-app
                Resolved decision is never lost; the extras are deleted."""
                ordered = sorted(
                    rows, key=lambda r: (0 if r.get("is_resolved") else 1, str(r["id"])))
                return ordered[0], ordered[1:]

            consumed: set[str] = set()   # existing ids already claimed this run
            inserts: list[dict[str, Any]] = []
            dup_ids: list[str] = []
            updated = adopted = collapsed = 0
            for r in m_rows:
                ext = r["external_id"]
                cands = [c for c in ext_to_rows.get(ext, [])
                         if str(c["id"]) not in consumed]
                # Transition: a row that just gained a MID won't match by the
                # new key yet — adopt the existing row(s) under the OLD derived
                # key, migrating to the stable MID (keeps is_resolved).
                if not cands:
                    cands = [c for c in ext_to_rows.get(r.get("_derived_id"), [])
                             if str(c["id"]) not in consumed]
                    if cands:
                        adopted += 1
                if cands:
                    keep, extras = _survivor(cands)
                    consumed.add(str(keep["id"]))
                    consumed.update(str(x["id"]) for x in extras)
                    # Update Excel-managed fields on the survivor; is_resolved
                    # (app-managed) survives. external_id migrates to the MID.
                    sb.table("meeting_logs").update({
                        "external_id":  ext,
                        "meeting_date": r["meeting_date"],
                        "donor_title":  r["donor_title"],
                        "rfp_uid":      r["rfp_uid"],
                        "remarks":      r["remarks"],
                        "actions":      r["actions"],
                        "owner":        r["owner"],
                        "deadline":     r["deadline"],
                    }).eq("id", keep["id"]).execute()
                    updated += 1
                    dup_ids.extend(str(x["id"]) for x in extras)  # same-meeting dupes
                    collapsed += len(extras)
                else:
                    inserts.append({k: v for k, v in r.items()
                                    if k != "_derived_id"})
            if inserts:
                sb.table("meeting_logs").insert(inserts).execute()
            # Remove the collapsed duplicate rows (in chunks).
            for _i in range(0, len(dup_ids), 100):
                sb.table("meeting_logs").delete().in_(
                    "id", dup_ids[_i:_i + 100]).execute()
            print(f"  meeting_logs: {updated} updated ({adopted} adopted->MID, "
                  f"{collapsed} duplicate(s) removed) · {len(inserts)} inserted "
                  f"(is_resolved preserved)")

    # --- engagement_logs — merge by external_id (avoid duplicate accumulation)
    print("[Engagement_Log -> engagement_logs]")
    e_rows = [r for row in _rows(wb["Engagement_Log"], header_row=6) if (r := map_engagement_row(row))]
    merge_by_external_id(
        "engagement_logs",
        e_rows,
        updatable_fields=[
            "engagement_date", "donor", "engagement_type", "format",
            "internal_lead", "donor_contacts", "purpose", "outcome", "linked_rfp_uid",
        ],
    )

    # --- active_grants — keyed on grant_id. Delete stale migration rows
    # FIRST (rows that existed in a prior sync but disappeared from Excel
    # — without this they linger forever and pollute the per-grant view).
    # App-added rows (source='app') are preserved.
    print("[Active_Grants_Log -> active_grants]")
    g_rows = [r for row in _rows(wb["Active_Grants_Log"], header_row=6) if (r := map_active_grant_row(row))]
    current_grant_ids = [r["grant_id"] for r in g_rows if r.get("grant_id")]
    if not dry_run:
        # Delete migration rows whose grant_id is no longer in Excel.
        existing_migration = (
            sb.table("active_grants")
            .select("grant_id")
            .eq("source", "migration")
            .execute()
            .data
            or []
        )
        existing_mig_ids = {r["grant_id"] for r in existing_migration if r.get("grant_id")}
        stale_ids = sorted(existing_mig_ids - set(current_grant_ids))
        if stale_ids:
            sb.table("active_grants").delete().in_("grant_id", stale_ids).execute()
            print(f"  active_grants: deleted {len(stale_ids)} stale migration row(s): {stale_ids}")
    upsert("active_grants", g_rows, conflict_key="grant_id")

    # --- narrative_logs — merge by external_id
    print("[Narrative_Log -> narrative_logs]")
    n_rows = [r for row in _rows(wb["Narrative_Log"], header_row=6) if (r := map_narrative_row(row))]
    merge_by_external_id(
        "narrative_logs",
        n_rows,
        updatable_fields=[
            "version_date", "narrative_title", "used_in", "used_with",
            "date_used", "status", "link_location", "owner",
        ],
    )

    # --- meeting_schedule
    print("[Schedule -> meeting_schedule]")
    s_rows = [r for row in _rows(wb["Schedule"], header_row=4) if (r := map_schedule_row(row))]
    upsert("meeting_schedule", s_rows, conflict_key="call_date")

    print("\nDone." + (" (dry run — no writes performed)" if dry_run else ""))


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    if not args.xlsx.exists():
        sys.exit(f"Excel file not found: {args.xlsx}")
    print(f"Reading: {args.xlsx}")
    # Tenant-aware sync: when launched from a tenant admin's "Sync Excel" (core.excel_sync
    # passes RFPIS_SYNC_TENANT_ID), stamp EVERY imported row to that tenant via the headless
    # override, so the workbook lands in the acting tenant's pipeline — not as NULL-tenant
    # rows hidden from everyone. Without the env var it behaves exactly as before.
    _tid = os.environ.get("RFPIS_SYNC_TENANT_ID")
    _tok = None
    if _tid and not args.dry_run:
        try:
            from auth.tenant_context import set_tenant_override
            _tok = set_tenant_override(_tid)
            print(f"Tenant-scoped import → tenant_id={_tid}")
        except Exception as _e:
            print(f"(tenant override unavailable, importing unscoped: {_e})")
    try:
        migrate(args.xlsx, dry_run=args.dry_run)
    finally:
        if _tok is not None:
            try:
                from auth.tenant_context import reset_tenant_override
                reset_tenant_override(_tok)
            except Exception:
                pass


if __name__ == "__main__":
    main()
